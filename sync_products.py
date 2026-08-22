"""
商品信息同步 — sync_products.py
================================
报活动前先抓取最新商品信息（后台接口方式，不模拟页面勾选）。

用法：
  python sync_products.py                # 默认抓 5 页（每页 50 商品）
  python sync_products.py --pages 3      # 抓 3 页（150 商品）
  python sync_products.py --pages 1-3    # 抓第 1~3 页（与核价 pages 区间格式一致）

流程：
  1. 连接现有 Edge（CDP 9222），复用登录态
  2. 调 searchForSemiSupplier 循环抓 N 页商品基础信息（SPU/SKC/SKU/颜色/尺码/货号）
  3. 逐 SKU 调 queryProductSkuPriceAndStatus 抓完整 17 站价格
  4. 生成 Excel + JSON 存 state/product_sync/（10 列：SPU ID/SKC ID/SKC货号/SKU ID/SKU货号/颜色属性/尺码属性/站点/活动申报价格/币种）

说明：
  - 纯读操作，不动任何页面状态（不开抽屉、不勾选、不提交）。
  - 通过页面 SDK fetch（自动附加 anti-content 签名）+ mallid 头（从 cookie 提取）。
"""
import sys
import os
import json
import time
import argparse
import datetime

# 项目根目录加入 sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from playwright.sync_api import sync_playwright

CDP_URL = "http://127.0.0.1:9222"
PAGE_URL_KEYWORD = "agentseller.temu.com/newon"
SEARCH_PATH = "/api/kiana/mms/robin/searchForSemiSupplier"
PRICE_PATH = "/api/kiana/mms/robin/queryProductSkuPriceAndStatus"
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "state", "product_sync")
BATCH = 30  # 价格接口并发批大小

# 同步结果（供 bridge 轮询读取）
RESULT_FILE = os.path.join(OUT_DIR, "latest_sync_result.json")


def log(msg):
    print(msg, flush=True)


def parse_pages_arg(text):
    """支持 '5' 或 '1-3' → 返回 (start_page, end_page)"""
    text = (text or "").strip()
    if not text:
        return 1, 5
    if "-" in text:
        a, b = text.split("-", 1)
        return int(a), int(b)
    n = int(text)
    return 1, n


def find_newon_page(browser):
    """在现有 Edge 标签页里找上新页，找不到则新建"""
    for ctx in browser.contexts:
        for page in ctx.pages:
            if PAGE_URL_KEYWORD in (page.url or ""):
                return page
    # 没有上新页：用第一个 context 新建
    ctx = browser.contexts[0]
    page = ctx.new_page()
    page.goto("https://agentseller.temu.com/newon/product-select", wait_until="domcontentloaded", timeout=60000)
    time.sleep(5)
    return page


def fetch_pages(page, start_page, end_page):
    """抓 N 页商品基础信息（searchForSemiSupplier）"""
    all_meta = []
    for pn in range(start_page, end_page + 1):
        log(f"  [分页] 抓第 {pn} 页…")
        js = """
        (async () => {
          const mallid = (document.cookie.match(/mallid=([^;]+)/) || [])[1] || '';
          const resp = await fetch(ARG_PATH, {
            method: 'POST', credentials: 'include',
            headers: {'Content-Type': 'application/json', 'mallid': mallid},
            body: JSON.stringify({pageSize: 50, pageNum: ARG_PAGE})
          });
          const j = await resp.json();
          const list = (j.result || {}).dataList || [];
          const out = [];
          for (const p of list) {
            const spuId = p.productId, goodsId = p.goodsId, spuName = p.productName;
            for (const skc of (p.skcList || [])) {
              const skcId = skc.skcId, color = skc.colorName || skc.color || '', skcCode = skc.extCode || '';
              const prv = (skc.supplierPriceReviewInfoList || [])[0] || {};
              for (const sku of (prv.productSkuList || [])) {
                const prop = (sku.productPropertyList || []).find(x => x.name === '尺码');
                out.push({
                  spuId, goodsId, spuName,
                  skcId, color, skcCode,
                  skuId: sku.skuId, skuCode: sku.extCode || '',
                  size: prop ? prop.value : ''
                });
              }
            }
          }
          return JSON.stringify(out);
        })()
        """
        result = page.evaluate(
            js.replace("ARG_PATH", json.dumps(SEARCH_PATH)).replace("ARG_PAGE", str(pn))
        )
        batch = json.loads(result)
        log(f"    {pn} 页: {len(batch)} 个 SKU")
        all_meta.extend(batch)
        time.sleep(0.5)
    # 去重（同一 SKU 可能跨页重复）
    seen = set()
    uniq = []
    for m in all_meta:
        if m["skuId"] not in seen:
            seen.add(m["skuId"])
            uniq.append(m)
    return uniq


def fetch_prices(page, meta_list):
    """逐 SKU 抓完整 17 站价格（queryProductSkuPriceAndStatus，页面内并发）"""
    price_map = {}
    total = len(meta_list)
    t0 = time.time()
    for i in range(0, total, BATCH):
        batch = meta_list[i:i + BATCH]
        js = """
        (async () => {
          const mallid = (document.cookie.match(/mallid=([^;]+)/) || [])[1] || '';
          const fetchOne = async (m) => {
            try {
              const resp = await fetch(ARG_PATH, {
                method: 'POST', credentials: 'include',
                headers: {'Content-Type': 'application/json', 'mallid': mallid},
                body: JSON.stringify({productSkuId: m.skuId})
              });
              const j = await resp.json();
              const sites = ((j.result || {}).siteSupplierPriceList) || [];
              return {skuId: m.skuId, sites: sites.map(s => ({
                siteName: s.siteName, price: s.supplierPriceValue, currency: s.supplierPriceCurrencyType
              }))};
            } catch (e) { return {skuId: m.skuId, sites: []}; }
          };
          const results = await Promise.all(ARG_BATCH.map(m => fetchOne(m)));
          return JSON.stringify(results);
        })()
        """
        result = page.evaluate(
            js.replace("ARG_PATH", json.dumps(PRICE_PATH)).replace("ARG_BATCH", json.dumps(batch))
        )
        try:
            for p in json.loads(result):
                price_map[p["skuId"]] = p["sites"]
        except Exception as e:
            log(f"    ⚠️ 批次解析失败: {e}")
        done = min(i + BATCH, total)
        if done % 300 < BATCH or done == total:
            el = time.time() - t0
            log(f"  [价格] {done}/{total} ({el:.0f}s)")
    return price_map


def build_rows(meta_list, price_map):
    """合并为基础 10 列 Excel 行"""
    rows = []
    for m in meta_list:
        sites = price_map.get(m["skuId"], [])
        if not sites:
            rows.append({
                "SPU ID": m["spuId"], "SKC ID": m["skcId"], "SKC货号": m["skcCode"],
                "SKU ID": m["skuId"], "SKU货号": m["skuCode"],
                "颜色属性": m["color"], "尺码属性": m["size"],
                "站点": "", "活动申报价格": "", "币种": ""
            })
            continue
        for s in sites:
            rows.append({
                "SPU ID": m["spuId"], "SKC ID": m["skcId"], "SKC货号": m["skcCode"],
                "SKU ID": m["skuId"], "SKU货号": m["skuCode"],
                "颜色属性": m["color"], "尺码属性": m["size"],
                "站点": s["siteName"],
                "活动申报价格": round(s["price"] / 100, 2) if s["price"] else "",
                "币种": s["currency"] or ""
            })
    return rows


def save_excel(rows, out_path):
    """写入 Excel（10 列 + 表头样式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"
    headers = ["SPU ID", "SKC ID", "SKC货号", "SKU ID", "SKU货号",
               "颜色属性", "尺码属性", "站点", "活动申报价格", "币种"]
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")
    for i, w in enumerate([12, 14, 12, 14, 26, 12, 10, 12, 14, 8], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    for cell in ws["I"][1:]:
        cell.number_format = "0.00"
    wb.save(out_path)


def write_result(payload):
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(RESULT_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main():
    ap = argparse.ArgumentParser(description="报活动前同步最新商品信息")
    ap.add_argument("--pages", default="5", help="页数：'5' 或 '1-3'（每页 50 商品，默认 5 页）")
    args = ap.parse_args()
    start_page, end_page = parse_pages_arg(args.pages)
    n_pages = end_page - start_page + 1

    t0 = time.time()
    log(f"== 商品信息同步 == 页范围 {start_page}-{end_page}（{n_pages} 页 × 50 商品）")
    log("[1/4] 连接 Edge (CDP 9222)…")

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(CDP_URL)
        page = find_newon_page(browser)
        log(f"  ✅ 已连接页面: {page.url[:80]}")

        log("[2/4] 抓取商品基础信息…")
        meta_list = fetch_pages(page, start_page, end_page)
        log(f"  ✅ 共 {len(meta_list)} 个 SKU（去重后）")

        log("[3/4] 抓取 17 站价格…")
        price_map = fetch_prices(page, meta_list)
        n_ok = sum(1 for s in price_map.values() if s)
        log(f"  ✅ 价格获取 {n_ok}/{len(meta_list)}")

        log("[4/4] 生成 Excel/JSON…")
        rows = build_rows(meta_list, price_map)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(OUT_DIR, exist_ok=True)
        excel_path = os.path.join(OUT_DIR, f"product_sync_{ts}.xlsx")
        json_path = os.path.join(OUT_DIR, f"product_sync_{ts}.json")
        save_excel(rows, excel_path)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=1)

        spus = len({r["SPU ID"] for r in rows})
        colors = len({(r["SPU ID"], r["SKC ID"]) for r in rows})
        skus = len({r["SKU ID"] for r in rows})
        elapsed = int(time.time() - t0)
        log(f"  ✅ Excel: {excel_path}")
        log(f"  ✅ JSON:  {json_path}")
        log(f"  ✅ 汇总: {spus} 商品 / {colors} 颜色 / {skus} SKU / {len(rows)} 行（{elapsed}s）")

        write_result({
            "ok": True,
            "pages": f"{start_page}-{end_page}",
            "spu_count": spus,
            "color_count": colors,
            "sku_count": skus,
            "row_count": len(rows),
            "elapsed_sec": elapsed,
            "excel_path": excel_path,
            "json_path": json_path,
            "excel_name": os.path.basename(excel_path),
            "finished_at": datetime.datetime.now().isoformat(),
        })


if __name__ == "__main__":
    main()
